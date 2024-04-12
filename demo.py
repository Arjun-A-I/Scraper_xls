import openpyxl
from googlesearch import search

# Path to the Excel file
path = "data.xlsx"

# Load the workbook
wb_obj = openpyxl.load_workbook(path)

# Get the active sheet
sheet_obj = wb_obj.active

# Create a list to store the startup names
startup_name = []

# Iterate through rows 2 to 46 (inclusive) and get the values from column 1
for row in range(2, 47):
    cell_obj = sheet_obj.cell(row=row, column=1)
    startup_name.append(cell_obj.value)

# Create a new workbook and worksheet to store the search results
result_wb = openpyxl.Workbook()
result_sheet = result_wb.active

# Iterate through the startup_name list and perform a Google search for each
for row, startup in enumerate(startup_name, start=1):
    query = startup
    for result in search(query, tld="co.in", num=1, stop=1, pause=0):
        # Add the startup name to the first column
        result_sheet.cell(row=row, column=1, value=startup)
        # Add the search result to the second column
        result_sheet.cell(row=row, column=2, value=result)

# Save the new workbook
result_wb.save("search_results.xlsx")

