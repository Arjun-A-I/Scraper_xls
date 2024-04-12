import instructor
import openai
from pydantic import BaseModel, Field
import json
import os
from dotenv import load_dotenv, find_dotenv
import requests
from bs4 import BeautifulSoup
import openpyxl

load_dotenv(find_dotenv())
api_key = os.getenv("OPENAI_API_KEY")

instructor_openai_client = instructor.patch(openai.Client(
    api_key=api_key, timeout=20000, max_retries=3
))

# Load the Excel file
workbook = openpyxl.load_workbook('search_results.xlsx')
sheet = workbook.active

class StartupData(BaseModel):
    startup_name: str = Field(description="Denotes the name of the startup/company.")
    website: str = Field(description="The website URL of the startup/company.")
    context_text: str = Field(description="A 20-word contextual hook that conveys how the SaaS tool 'Root' can be useful to the company and its product.")

# Create a new worksheet to store the results
result_worksheet = workbook.create_sheet("Results")
result_worksheet["A1"] = "Startup Name"
result_worksheet["B1"] = "Website"
result_worksheet["C1"] = "Contextual Hook"

# Iterate through rows 2 to 46 and process each URL
for row in range(2, 47):
    startup_name = sheet.cell(row=row, column=1).value
    website = sheet.cell(row=row, column=2).value

    if website and isinstance(website, str) and website.strip():
        try:
            print("Doing rn")
            response = requests.get(website)
            soup = BeautifulSoup(response.content, "html.parser")
            text = soup.get_text()
            text = text.replace("\n", "")
            text = text.replace("\t", "")

            # Split the text into smaller chunks
            chunks = [text[i:i+4096] for i in range(0, len(text), 4096)]

            for chunk in chunks:
                completion = instructor_openai_client.chat.completions.create(
                    model="gpt-4-turbo-2024-04-09",
                    messages=[
                        {
                            "role": "user",
                            "content": f"You are an expert content writer and company analyst at Root, a SaaS tool that provides a personalized experience to users through quizzes, recommendations, and a personalized UI. Write a 20-word contextual hook for the following company details: {chunk}"
                        }
                    ],
                    response_model=StartupData
                )

                context_text = completion.model_dump().get("context_text")
                result_worksheet.append([startup_name, website, context_text])

            # Save the workbook after processing each row
            workbook.save("results.xlsx")
        except (requests.exceptions.RequestException, openai.BadRequestError) as e:
            print(f"Error processing website {website}: {e}")
            # Save the workbook before continuing to the next row
            workbook.save("results.xlsx")
    else:
        print(f"Skipping row {row} because the website is empty or not a valid string.")

# Save the workbook one final time
workbook.save("results.xlsx")